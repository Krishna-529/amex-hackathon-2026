import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

PATH = 'ZKD-Feature-Checklist.xlsx'
wb = openpyxl.load_workbook(PATH)

if 'Member scenarios' in wb.sheetnames:
    del wb['Member scenarios']
ws = wb.create_sheet('Member scenarios')

GREEN = PatternFill('solid', fgColor='C6EFCE')
AMBER = PatternFill('solid', fgColor='FFEB9C')
RED = PatternFill('solid', fgColor='FFC7CE')
HEADER_FILL = PatternFill('solid', fgColor='1F3864')
HEADER_FONT = Font(bold=True, color='FFFFFF', size=11)
TITLE_FONT = Font(bold=True, size=13, color='1F3864')
WRAP = Alignment(wrap_text=True, vertical='top')
THIN = Border(*[Side(style='thin', color='BFBFBF')]*4)

ws.append(['Member scenarios - real-life check (generated 2026-08-18)'])
ws['A1'].font = TITLE_FONT
ws.append(['Think like an actual Amex card member mid-trip: what can go wrong, and does ZKD Concierge have an answer?'])
ws['A2'].font = Font(italic=True, color='595959')
ws.append([])

header = ['#', 'Stage', 'Real-life scenario', 'What the member needs', 'ZKD answer?',
          'Where it lives in code', 'How to see it with your own eyes', 'Status', 'Notes']
ws.append(header)
for c, v in enumerate(header, 1):
    cell = ws.cell(row=3, column=c)
    cell.fill = HEADER_FILL
    cell.font = HEADER_FONT
    cell.alignment = Alignment(wrap_text=True, vertical='center', horizontal='center')
    cell.border = THIN

rows = [
    # (stage, scenario, need, answer, code, verify, status, notes)
    ('Before departure',
     'Flight is cancelled a few hours before departure',
     'Someone rebooks me onto the best alternative without me having to spend the day on hold',
     'Covered - detectDisruption() starts the pipeline: search alternatives -> rank against my rules -> consent window -> saga books -> journal',
     'server/engine/simulation.ts, server/pipeline/',
     'Login as priya@zkd.demo, open /ops, trigger a disruption on a flight you are booked on, then open /recovery/[id]',
     'Covered', 'Detection itself is a manual /ops button - a real airline feed was never wired (documented gap, row 40 of the feature sheet)'),
    ('Before departure',
     'The flight is delayed enough that I will miss my connection',
     'Re-time my whole trip - keep the onward leg working, move the hotel and cab with it',
     'Covered - connecting itinerary exists (MAA->DEL->LHR, one PNR), deadlines resolve in destination local time; re-timed resolution keeps the original ticket',
     'server/domain/seed.ts (u1/u2), server/deadline.ts, simulation.ts finalizeResolution(re-timed)',
     'Trigger a disruption on u1 in /ops and watch /recovery show the LHR leg protected',
     'Covered', 'connectionRisk is null in forecasts - real rotation/connection linkage (OAG Connections) still pending production approval'),
    ('Before departure',
     'The airline overbooks / denies boarding at the gate',
     'Get the next seat and the airline compensation without arguing at the gate',
     'Not covered - no gate/status feed and no denied-boarding handling; only cancellations and pre-departure rebooking exist',
     'n/a',
     'Search the repo for "denied" / "overbook" - nothing',
     'Not covered', 'Would need a live status/boarding feed and a claims step the pipeline does not have'),
    ('Before departure',
     'Flight diverted or turned back mid-air',
     'Know where the plane is going and have a plan for wherever I land',
     'Not covered - documented out of scope',
     'context.md known limitations',
     'Read context.md "Known limitations"',
     'Not covered', 'Deliberate scope cut - diversions/turn-backs out of scope'),
    ('Before departure',
     'It is a storm; will my flight really go?',
     'An honest risk read tied to the actual weather',
     'Partial - weather is fetched (Open-Meteo/NOAA) but only feeds the explanation narrative; the trained model has NO weather feature in v1',
     'server/weather.ts, riskModel.ts header "KNOWN v1 GAP"',
     'Open /flights/[id] and read the explanation; the weather sentence appears, the score never changes because of it',
     'Partial', 'Weather-as-feature is the clearly-scoped next retrain (NOAA ISD join) - stated in the code'),
    ('During disruption',
     'The alternative I was shown vanishes while I am deciding',
     'Get the next option that still works - not a "sorry" page',
     'Covered - revalidateChoice() re-checks the chosen alt against live inventory at the moment of spend and cascades to the next viable one',
     'server/engine/simulation.ts revalidateChoice()',
     'Hard to force live; the code path is exercised by the verify suite',
     'Covered', 'Only a confirmed "gone" triggers a switch; uncheckable candidates are left alone'),
    ('During disruption',
     'I am travelling with my whole family - six of us',
     'Never put us on two different flights',
     'Covered - altsForParty() removes anything that cannot seat the PNR size, and "we will not split your party" is enforced server-side, not just in UI',
     'server/domain/altsForParty.ts, simulation.ts resolveTask(choose)',
     'Login as arjun@zkd.demo (party of 6) on f-multi and compare options vs rohan@zkd.demo (party of 2)',
     'Covered', 'This is the reason two option classes exist (carrier-protected vs market)'),
    ('During disruption',
     'I am asleep / offline / just not answering',
     'My standing choice should decide, safely',
     'Covered - autopilot books immediately (re-validating inventory first); ask-me-first: free fixes auto-book, anything costing money hands over',
     'simulation.ts createTaskForBooking (autopilot branch) + settleExpired()',
     'Set /settings to each mode, trigger a disruption, and do not touch the recovery page',
     'Covered', 'A real behavioural branch in the engine, not UI copy'),
    ('During disruption',
     'The proposed fix would blow past what my card will authorise',
     'Nobody charges the card; stop and tell me',
     'Covered - per-transaction cap (MyCa, Rs 25,000) is enforced before ANY spend path: approve, autopilot, and silent timeout all check it',
     'server/myca.ts, simulation.ts (overCap branches)',
     'Trigger a recovery whose plan costs more than the cap and read the note',
     'Covered', 'The cap is the card limit, not a preference - even explicit approval cannot pass it'),
    ('On the ground',
     'I land with no hotel and no plan',
     'A room near the airport that fits the airline\'s duty-of-care rate',
     'Covered (logic) / Partial (data) - real LiteAPI search happens, but the sandbox tier often returns exactly ONE hotel per city, and the fallback rows are generated',
     'server/hotels/, server/engine/groundCache.ts, server/liteapi.ts',
     'GET /api/hotels?iata=DEL -> often a single result; compare to the mockHotels fallback rows',
     'Partial', 'Search proved live today: "Kesarval The Fern Goa" and "Hotel A-10 International" came back real'),
    ('On the ground',
     'Cab from the airport',
     'A car that fits the party, within the transfer allowance',
     'Partial - cab list is always generated (server/mockCabs.ts); there is no real ground-transfer supplier in the build',
     'server/mockCabs.ts, groundCache.ts',
     'Read the groundCache header comment - it says so in plain words',
     'Partial', 'Deliberate: "no real ground-transfer supplier to call"'),
    ('On the ground',
     'Baggage lost or delayed',
     'Track it, get reimbursed, get essentials',
     'Not covered - no baggage feature anywhere',
     'n/a',
     'Search the repo for baggage - nothing',
     'Not covered', 'Out of product scope; would be a separate service'),
    ('On the ground',
     'Medical problem mid-trip',
     'Help with care, not just transport',
     'Not covered - no health/duty-of-care escalation beyond hotel+flight',
     'n/a',
     'n/a',
     'Not covered', 'Out of scope'),
    ('Money & claims',
     'The airline owes me money (cancellation compensation, hotel bill)',
     'File the claim / get the money without paperwork',
     'Partial - the pipeline knows what the airline "owes" (carrier-protected alts at fare 0, owedNow), but nothing actually files a claim or issues a refund; the original ticket is never cancelled',
     'server/domain/pricing.ts, pipeline/saga.ts',
     'Read the saga commit records - they record intent:<ref>, not ticketing',
     'Partial', 'The claim/duty-of-care step is narrative, not a real process'),
    ('Money & claims',
     'I paid for the original ticket - who refunds it?',
     'The old ticket is cancelled and refunded, the new one issued',
     'Not covered - no original-ticket cancellation path; the saga books intent on the new option and leaves the old PNR untouched',
     'pipeline/saga.ts (intent-only commits)',
     'grep PIPELINE_ALLOW_MUTATIONS in server/',
     'Not covered', 'Row 27 of the feature sheet: "Nothing can actually buy a ticket" - applies symmetrically to refunds'),
    ('Money & claims',
     'The price on the offer is not my currency',
     'See a price I can compare, never a silently-converted one',
     'Covered - needsConversion() refuses to convert; offers in a foreign currency are shown but excluded with an explicit reason',
     'server/domain/altsFromOffers.ts, hotels/types.ts',
     'Today\'s live result: ~30 real Duffel EUR offers all marked "Priced in EUR while your..." - the guard works',
     'Covered', 'Trade-off: today the only bookable market options are the mock/travelport rows'),
    ('Communication',
     'I never got the alert',
     'Know that I was told - or that nobody could reach me',
     'Partial - every attempt is ledger-logged (delivered vs skipped), but WhatsApp is BLOCKED on the trial Twilio account and Android push is untested',
     'server/notify/, server/decisionLedger.ts',
     'cat zkd-app/server/.state/notifications.jsonl (if present)',
     'Partial', 'SEE BUG NOTE BELOW - the ledger files could not be found in the running instance'),
    ('Communication',
     'I am on my phone, not my laptop',
     'The full flow on Android',
     'Partial - Android app exists (4 screens) but has no pre-authorisation flow and no consent-settings screen; remote push needs an EAS projectId that app.json does not have',
     'zkd-android/',
     'Run npx expo run:android and compare screens to the web app',
     'Partial', 'Documented subset'),
    ('Devices & identity',
     'I book for my mother who is not a card member',
     'Manage her trip and her recovery on her behalf',
     'Partial - companions can ride on the member\'s PNR as Traveller records (no login), but the "card-member-books-for-someone-else" consent model is an unresolved mentor question',
     'server/domain/seed.ts (companions), context.md',
     'Read context.md known limitations',
     'Partial', 'Documented as unresolved'),
    ('Trust',
     '"80%"? Really?',
     'Understand what the number means and why',
     'Covered - the gauge is a 0-100 percentile, the real calibrated probability (~2-8%) is shown beside it, and the SHAP audit panel labels every input as real / synthetic-market-estimate / population-average',
     'components/ForecastAudit.tsx, app/flights/page.tsx',
     'Open /flights/[id] and read the audit panel + the note under the gauge',
     'Covered', 'The system is deliberately honest about cold-starting Indian carriers'),
    ('Audit',
     'You told me this was going to happen - prove it',
     'A checkable record of every prediction and every alert',
     'Covered (code) / UNVERIFIED (runtime) - logPrediction/logThresholdEvaluation/logNotification run inside applyScore, but no .state/*.jsonl files exist under zkd-app/server in the running instance',
     'server/decisionLedger.ts',
     'ls zkd-app/server/.state/ - currently EMPTY despite live forecasts',
     'Partial', 'POSSIBLE BUG: if the dev server is started with a different cwd, ledger writes land somewhere else - worth a targeted check before the demo'),
    ('Money & claims',
     'I am worried the recovery could put me in negative balance / out of pocket',
     'Know that everything is charged to my Amex and anything refunded comes back to that same card - and that the system stops before spending past my cap',
     'Covered - payment routing invariant: all spend on the member\'s Amex card, all refunds to the same card; per-transaction cap checked on every spend path (approve, autopilot, silent timeout), so the card can never go negative from this flow',
     'documentation/design/03-action-policy.md §10 (new 2026-08-18), server/engine/simulation.ts overCap branches',
     'Trigger a recovery over the Rs 25,000 cap on /ops and read the explanation',
     'Covered', 'Policy written up as §10; cap enforcement is real, refunds still mocked end-to-end'),
]

status_fill = {'Covered': GREEN, 'Partial': AMBER, 'Not covered': RED}
r = 4
for i, (stage, scenario, need, answer, code, verify, status, notes) in enumerate(rows, 1):
    vals = [i, stage, scenario, need, answer, code, verify, status, notes]
    ws.append(vals)
    for c in range(1, len(header) + 1):
        cell = ws.cell(row=r, column=c)
        cell.alignment = WRAP
        cell.border = THIN
        if c == 8:
            cell.fill = status_fill[status]
            cell.font = Font(bold=True)
    r += 1

ws.append([])
summary = ws.max_row + 1
ws.cell(row=summary, column=1, value='Summary').font = Font(bold=True)
ws.append(['', '', 'Covered:', sum(1 for x in rows if x[6] == 'Covered'), 'Partial:', sum(1 for x in rows if x[6] == 'Partial'), 'Not covered:', sum(1 for x in rows if x[6] == 'Not covered')])

widths = [4, 16, 38, 38, 52, 34, 40, 13, 46]
for c, w in enumerate(widths, 1):
    ws.column_dimensions[get_column_letter(c)].width = w
ws.freeze_panes = 'A4'

wb.save(PATH)
print('saved. sheets:', wb.sheetnames)
